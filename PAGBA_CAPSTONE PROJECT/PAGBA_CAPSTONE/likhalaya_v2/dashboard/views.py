import json
import csv
import functools
from io import BytesIO
from datetime import timedelta, date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import zipfile

from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, Avg, Case, When, Value, IntegerField
from django.db.models.functions import TruncMonth, TruncDate
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator

from store.models import Product, Category, ContactMessage, LivelihoodVideo, ProductImage
from store.forms import ProductForm, ProductImageFormSet
from orders.models import Order, OrderItem, Notification
from accounts.models import CustomUser, ActivityLog
from accounts.activity import log_activity


# ─── Decorators ────────────────────────────────────────────────────────────────
def staff_required(view_func):
    @functools.wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_staff_user():
            messages.error(request, 'Access denied. Staff only.')
            return redirect('store:home')
        return view_func(request, *args, **kwargs)
    return wrapper


def admin_required(view_func):
    @functools.wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_admin_user():
            messages.error(request, 'Access denied. Admins only.')
            return redirect('dashboard:home')
        return view_func(request, *args, **kwargs)
    return wrapper


def _xlsx_bytes(wb):
    """
    Save an openpyxl Workbook to bytes with [Content_Types].xml and _rels/.rels
    placed first in the zip archive. openpyxl doesn't guarantee this ordering,
    and while Microsoft Excel is lenient about it, WPS Office's parser is not —
    files that open fine in Excel can fail to open ("Unable to open file") in
    WPS unless the standard OPC entry order is respected. Re-packaging the zip
    this way makes the export open cleanly in both.
    """
    raw = BytesIO()
    wb.save(raw)
    raw.seek(0)

    src = zipfile.ZipFile(raw)
    names = src.namelist()
    priority = ['[Content_Types].xml', '_rels/.rels', 'xl/workbook.xml', 'xl/_rels/workbook.xml.rels']
    ordered = [n for n in priority if n in names] + [n for n in names if n not in priority]

    out = BytesIO()
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as z:
        for n in ordered:
            z.writestr(n, src.read(n))
    out.seek(0)
    return out.getvalue()


# ─── Dashboard Home ─────────────────────────────────────────────────────────────
@staff_required
def dashboard_home(request):
    now = timezone.now()
    thirty_days_ago = now - timedelta(days=30)
    seven_days_ago = now - timedelta(days=7)

    # ── True calendar-month boundaries (not a rolling 30-day window) ──
    # "This month" = 1st of current month through now.
    # "Last month" = 1st through last day of the previous month.
    # Nothing is ever deleted — these are just date filters on the same Order table,
    # so the numbers naturally roll over to ₱0 on the 1st of a new month.
    this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_end = this_month_start - timedelta(microseconds=1)
    last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    this_month_revenue = Order.objects.filter(
        status='delivered', created_at__gte=this_month_start
    ).aggregate(t=Sum('total'))['t'] or 0
    last_month_revenue = Order.objects.filter(
        status='delivered', created_at__gte=last_month_start, created_at__lt=this_month_start
    ).aggregate(t=Sum('total'))['t'] or 0

    # % change vs last month, for a "up/down from last month" indicator on the card
    if last_month_revenue:
        revenue_change_pct = round(((this_month_revenue - last_month_revenue) / last_month_revenue) * 100, 1)
    else:
        revenue_change_pct = 100.0 if this_month_revenue else 0.0

    stats = {
        'total_orders': Order.objects.count(),
        'orders_today': Order.objects.filter(created_at__date=now.date()).count(),
        'total_revenue': Order.objects.filter(status='delivered').aggregate(t=Sum('total'))['t'] or 0,
        'monthly_revenue': this_month_revenue,
        'last_month_revenue': last_month_revenue,
        'revenue_change_pct': revenue_change_pct,
        'pending_orders': Order.objects.filter(status='pending').count(),
        'confirmed_orders': Order.objects.filter(status='confirmed').count(),
        'delivered_orders': Order.objects.filter(status='delivered').count(),
        'cancelled_orders': Order.objects.filter(status='cancelled').count(),
        'total_products': Product.objects.filter(is_active=True).count(),
        'low_stock': Product.objects.filter(stock__lte=5, is_active=True).count(),
        'out_of_stock': Product.objects.filter(stock=0, is_active=True).count(),
        'total_customers': CustomUser.objects.filter(role='customer').count(),
        'total_staff': CustomUser.objects.filter(role__in=['staff', 'coordinator']).count(),
        'new_customers': CustomUser.objects.filter(role='customer', created_at__gte=thirty_days_ago).count(),
        'total_messages': ContactMessage.objects.count(),
        'unread_messages': ContactMessage.objects.filter(is_read=False).count(),
        'total_categories': Category.objects.filter(is_active=True).count(),
    }

    # Chart 1: last 7 days orders (trend — is business picking up?)
    daily_orders = []
    daily_labels = []
    for i in range(6, -1, -1):
        d = (now - timedelta(days=i)).date()
        cnt = Order.objects.filter(created_at__date=d).count()
        daily_orders.append(cnt)
        daily_labels.append(d.strftime('%b %d'))

    # Chart 2: order status distribution (only statuses that actually have orders,
    # so the legend/chart isn't cluttered with zero-count slices)
    status_display = dict(Order.STATUS_CHOICES)
    status_colors = {
        'pending': '#f39c12', 'confirmed': '#3498db', 'processing': '#1abc9c',
        'shipped': '#9b59b6', 'delivered': '#27ae60', 'cancelled': '#e74c3c',
    }
    status_counts = {s[0]: Order.objects.filter(status=s[0]).count() for s in Order.STATUS_CHOICES}
    status_counts = {k: v for k, v in status_counts.items() if v > 0}

    recent_orders = Order.objects.select_related('user').prefetch_related('items').order_by('-created_at')[:8]
    low_stock_products = Product.objects.filter(stock__lte=5, is_active=True).select_related('category')[:5]
    recent_customers = CustomUser.objects.filter(role='customer').order_by('-created_at')[:5]

    ctx = {
        'stats': stats,
        'recent_orders': recent_orders,
        'low_stock_products': low_stock_products,
        'recent_customers': recent_customers,
        'chart_daily_labels': json.dumps(daily_labels),
        'chart_daily_orders': json.dumps(daily_orders),
        'chart_status_labels': json.dumps([status_display[k] for k in status_counts.keys()]),
        'chart_status_data': json.dumps(list(status_counts.values())),
        'chart_status_colors': json.dumps([status_colors[k] for k in status_counts.keys()]),
    }
    return render(request, 'dashboard/home.html', ctx)


# ─── Messages ───────────────────────────────────────────────────────────────────
@staff_required
def message_list(request):
    messages_qs = ContactMessage.objects.order_by('-created_at')
    read_filter = request.GET.get('read', '')
    search = request.GET.get('q', '')
    if read_filter in {'true', 'false'}:
        messages_qs = messages_qs.filter(is_read=(read_filter == 'true'))
    if search:
        messages_qs = messages_qs.filter(Q(name__icontains=search) | Q(subject__icontains=search) | Q(email__icontains=search))
    paginator = Paginator(messages_qs, 15)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'dashboard/messages/list.html', {
        'messages_list': page_obj,
        'page_obj': page_obj,
        'read_filter': read_filter,
        'search': search,
    })


@staff_required
def message_detail(request, pk):
    msg = get_object_or_404(ContactMessage, pk=pk)
    if not msg.is_read:
        msg.is_read = True
        msg.save(update_fields=['is_read'])
    return render(request, 'dashboard/messages/detail.html', {'message': msg})


@staff_required
def message_delete(request, pk):
    msg = get_object_or_404(ContactMessage, pk=pk)
    if request.method == 'POST':
        msg.delete()
        messages.success(request, 'Message deleted.')
        return redirect('dashboard:message_list')
    return redirect('dashboard:message_detail', pk=pk)


# ─── Products ───────────────────────────────────────────────────────────────────
@staff_required
def product_list(request):
    products = Product.objects.select_related('category').order_by('-created_at')
    q = request.GET.get('q', '')
    cat_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')
    if q:
        products = products.filter(Q(name__icontains=q) | Q(artisan_name__icontains=q))
    if cat_filter:
        products = products.filter(category__id=cat_filter)
    if status_filter == 'active':
        products = products.filter(is_active=True)
    elif status_filter == 'inactive':
        products = products.filter(is_active=False)
    elif status_filter == 'low_stock':
        products = products.filter(stock__lte=5)
    paginator = Paginator(products, 15)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'dashboard/products/list.html', {
        'products': page_obj,
        'page_obj': page_obj,
        'q': q,
        'categories': Category.objects.filter(is_active=True),
        'cat_filter': cat_filter,
        'status_filter': status_filter,
    })


@staff_required
def product_detail(request, pk):
    product = get_object_or_404(
        Product.objects.select_related('category').prefetch_related('extra_images'), pk=pk
    )
    gallery_by_size = {'General': [], 'Small': [], 'Medium': [], 'Large': []}
    size_labels = {'S': 'Small', 'M': 'Medium', 'L': 'Large'}
    for img in product.extra_images.all():
        label = size_labels.get(img.size, 'General')
        gallery_by_size[label].append(img)
    return render(request, 'dashboard/products/detail.html', {
        'product': product,
        'gallery_by_size': gallery_by_size,
    })


@staff_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        image_formset = ProductImageFormSet(request.POST, request.FILES, prefix='images')
        if form.is_valid() and image_formset.is_valid():
            product = form.save()
            image_formset.instance = product
            image_formset.save()
            log_activity(request, 'create', f'Created product "{product.name}"',
                         resource='Product', resource_label=product.name,
                         new_value=f'{product.price_display} · stock {product.stock}')
            messages.success(request, f'Product "{product.name}" created successfully!')
            return redirect('dashboard:product_list')
    else:
        form = ProductForm()
        image_formset = ProductImageFormSet(prefix='images')
    return render(request, 'dashboard/products/form.html', {
        'form': form, 'image_formset': image_formset, 'title': 'Add Product',
    })


@staff_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        # Snapshot every trackable field before the form overwrites them,
        # so the activity log reports whichever ones actually changed
        # (not just price/stock — description, active status,
        # category, artisan, and image all count too).
        old_values = {
            'name': product.name,
            'description': product.description,
            'category': product.category.name if product.category else '(none)',
            'price_min': product.price_min,
            'price_medium': product.price_medium,
            'price_max': product.price_max,
            'stock': product.stock,
            'is_active': product.is_active,
            'artisan_name': product.artisan_name,
            'image': product.image.name if product.image else '',
        }
        form = ProductForm(request.POST, request.FILES, instance=product)
        image_formset = ProductImageFormSet(request.POST, request.FILES, instance=product, prefix='images')
        if form.is_valid() and image_formset.is_valid():
            form.save()
            image_formset.save()

            new_values = {
                'name': product.name,
                'description': product.description,
                'category': product.category.name if product.category else '(none)',
                'price_min': product.price_min,
                'price_medium': product.price_medium,
                'price_max': product.price_max,
                'stock': product.stock,
                'is_active': product.is_active,
                    'artisan_name': product.artisan_name,
                'image': product.image.name if product.image else '',
            }

            field_labels = {
                'name': 'Name',
                'description': 'Description',
                'category': 'Category',
                'price_min': 'Min price',
                'price_medium': 'Medium price',
                'price_max': 'Max price',
                'stock': 'Stock',
                'is_active': 'Active status',
                'artisan_name': 'Artisan',
                'image': 'Image',
            }

            changes = []
            for field, old_val in old_values.items():
                new_val = new_values[field]
                if old_val != new_val:
                    if field == 'image':
                        old_display = old_val.rsplit('/', 1)[-1] if old_val else '(none)'
                        new_display = new_val.rsplit('/', 1)[-1] if new_val else '(none)'
                    elif field == 'description':
                        # Full text is kept here; the detail page truncates
                        # long values visually and offers a "see all" toggle.
                        old_display = old_val or '(empty)'
                        new_display = new_val or '(empty)'
                    elif field == 'is_active':
                        old_display = 'On' if old_val else 'Off'
                        new_display = 'On' if new_val else 'Off'
                    else:
                        old_display = old_val
                        new_display = new_val
                    changes.append(f'{field_labels[field]}: {old_display} → {new_display}')

            previous_value = '; '.join(changes) if changes else 'No fields changed'
            new_value = f'{len(changes)} field(s) updated' if changes else 'No fields changed'

            log_activity(request, 'update', f'Updated product "{product.name}"',
                         resource='Product', resource_label=product.name,
                         previous_value=previous_value,
                         new_value=new_value)
            messages.success(request, f'Product "{product.name}" updated!')
            return redirect('dashboard:product_list')
    else:
        form = ProductForm(instance=product)
        image_formset = ProductImageFormSet(instance=product, prefix='images')
    return render(request, 'dashboard/products/form.html', {
        'form': form, 'image_formset': image_formset, 'title': 'Edit Product', 'product': product,
    })


@admin_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        name = product.name
        old_snapshot = f'{product.price_display} · stock {product.stock}'
        product.archive(by_user=request.user)
        log_activity(request, 'delete', f'Archived product "{name}"',
                     resource='Product', resource_label=name,
                     previous_value=old_snapshot)
        messages.success(request, f'Product "{name}" moved to Archive. You can restore it anytime.')
        return redirect('dashboard:product_list')
    return render(request, 'dashboard/products/confirm_delete.html', {'product': product})


@admin_required
def product_restore(request, pk):
    product = get_object_or_404(Product.all_objects, pk=pk, is_deleted=True)
    if request.method == 'POST':
        product.restore()
        log_activity(request, 'restore', f'Restored product "{product.name}"',
                     resource='Product', resource_label=product.name)
        messages.success(request, f'Product "{product.name}" restored.')
    return redirect('dashboard:archive')


# ─── Orders ─────────────────────────────────────────────────────────────────────
@staff_required
def order_list(request):
    orders = Order.objects.select_related('user').prefetch_related('items').order_by('-created_at')
    status = request.GET.get('status', '')
    search = request.GET.get('q', '')
    payment = request.GET.get('payment', '')
    if status:
        orders = orders.filter(status=status)
    if search:
        orders = orders.filter(Q(full_name__icontains=search) | Q(email__icontains=search) | Q(pk__icontains=search))
    if payment:
        orders = orders.filter(payment_method=payment)
    paginator = Paginator(orders, 15)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'dashboard/orders/list.html', {
        'orders': page_obj,
        'page_obj': page_obj,
        'status_filter': status,
        'search': search,
        'payment_filter': payment,
        'status_choices': Order.STATUS_CHOICES,
        'payment_choices': Order.PAYMENT_CHOICES,
    })


@staff_required
def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Order.STATUS_CHOICES):
            old_status = order.status
            if new_status == 'cancelled' and old_status != 'cancelled':
                order.previous_status = old_status
            order.status = new_status
            order.save()

            if new_status == 'cancelled' and old_status != 'cancelled':
                order.restock_items()

            if new_status != old_status:
                log_activity(
                    request, 'status_change',
                    f'Order {order.order_number} status: {old_status} → {new_status}',
                    resource='Order', resource_label=order.order_number,
                    previous_value=old_status, new_value=new_status,
                )

            if new_status != old_status and order.user:
                status_messages = {
                    'confirmed': f'Your order {order.order_number} has been confirmed.',
                    'shipped': f'Good news! Your order {order.order_number} has been shipped and is on its way.',
                    'delivered': f'Your order {order.order_number} has been delivered. Enjoy!',
                }
                note_message = status_messages.get(new_status)
                if note_message:
                    Notification.objects.create(user=order.user, order=order, message=note_message)

            messages.success(request, f'Order {order.order_number} updated: {old_status} → {new_status}.')
            return redirect('dashboard:order_detail', pk=pk)
    return render(request, 'dashboard/orders/detail.html', {
        'order': order,
        'status_choices': Order.STATUS_CHOICES,
    })


@staff_required
def order_export_csv(request):
    """Branded Excel export of the Orders list — same LIKHALAYA letterhead
    look as the Sales Report export (navy/teal/gold palette)."""
    orders = Order.objects.prefetch_related('items').order_by('-created_at')
    now = timezone.now()

    status_colors = {
        'pending':    'D0884E',
        'confirmed':  '4fa9e8',
        'processing': '7c8598',
        'shipped':    '4fd1c5',
        'delivered':  '3ecf8e',
        'cancelled':  'ff6b6b',
    }

    # ── Styling constants (mirrors the dashboard's teal/gold/navy palette) ──
    NAVY = '0e1420'
    TEAL = '4fd1c5'
    GOLD = 'D0884E'
    LIGHT_GRAY = 'F4F6F8'
    WHITE = 'FFFFFF'
    ZEBRA = 'FAFBFC'

    title_font = Font(name='Calibri', size=18, bold=True, color=NAVY)
    subtitle_font = Font(name='Calibri', size=10, italic=True, color='7C8598')
    header_font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    body_font = Font(name='Calibri', size=10, color='333333')
    badge_font = Font(name='Calibri', size=9, bold=True, color=WHITE)

    header_fill = PatternFill('solid', fgColor=TEAL)
    card_fill = PatternFill('solid', fgColor=LIGHT_GRAY)

    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Orders'
    ws.sheet_view.showGridLines = False

    columns = [
        ('Order #', 14), ('Customer', 20), ('Email', 26), ('Phone', 14),
        ('Status', 13), ('Payment', 16), ('Subtotal', 12), ('Shipping', 12),
        ('Total', 12), ('Date', 18),
    ]
    last_col_letter = get_column_letter(len(columns))
    for i, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    row = 1

    # ── Letterhead ──
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    c = ws[f'A{row}']
    c.value = 'LIKHALAYA'
    c.font = title_font
    c.fill = card_fill
    c.alignment = center
    ws.row_dimensions[row].height = 26
    row += 1

    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    c = ws[f'A{row}']
    c.value = 'PDL Market — Orders Export'
    c.font = Font(name='Calibri', size=11, bold=True, color=GOLD)
    c.fill = card_fill
    c.alignment = center
    row += 1

    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    c = ws[f'A{row}']
    c.value = f'Generated {now.strftime("%B %d, %Y at %I:%M %p")} — {orders.count()} orders'
    c.font = subtitle_font
    c.fill = card_fill
    c.alignment = center
    row += 2

    header_row = row
    for col_idx, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[header_row].height = 20
    ws.freeze_panes = f'A{header_row + 1}'
    row = header_row + 1

    for i, o in enumerate(orders):
        fill = PatternFill('solid', fgColor=ZEBRA) if i % 2 else PatternFill('solid', fgColor=WHITE)
        values = [
            o.order_number, o.full_name, o.email, o.phone,
            None,  # status handled separately as a colored badge
            o.get_payment_method_display(),
            float(o.subtotal), float(o.shipping_fee), float(o.total),
            timezone.localtime(o.created_at).replace(tzinfo=None),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx)
            if col_idx == 5:
                cell.value = o.get_status_display().upper()
                cell.font = badge_font
                cell.fill = PatternFill('solid', fgColor=status_colors.get(o.status, '888888'))
                cell.alignment = center
                cell.border = border
                continue
            cell.value = val
            cell.font = body_font
            cell.fill = fill
            cell.border = border
            if col_idx in (7, 8, 9):
                cell.number_format = '"₱"#,##0.00'
            elif col_idx == 10:
                cell.number_format = 'yyyy-mm-dd hh:mm'
            cell.alignment = center
        row += 1

    ws.auto_filter.ref = f'A{header_row}:{last_col_letter}{row - 1}'

    row += 1
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    footer = ws[f'A{row}']
    footer.value = 'Likhalaya PDL Market — Confidential internal orders export'
    footer.font = Font(name='Calibri', size=8, italic=True, color='AAAAAA')
    footer.alignment = center

    ws.print_area = f'A1:{last_col_letter}{row}'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buffer_bytes = _xlsx_bytes(wb)
    response = HttpResponse(
        buffer_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Likhalaya_Orders_{now.strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Users ──────────────────────────────────────────────────────────────────────
@admin_required
def user_list(request):
    role_order = Case(
        When(role='admin', then=Value(0)),
        When(role='staff', then=Value(1)),
        When(role='coordinator', then=Value(1)),
        When(role='customer', then=Value(2)),
        default=Value(3),
        output_field=IntegerField(),
    )
    users = CustomUser.objects.annotate(role_rank=role_order).order_by('role_rank', '-created_at')
    search = request.GET.get('q', '')
    role_filter = request.GET.get('role', '')
    if search:
        users = users.filter(Q(username__icontains=search) | Q(first_name__icontains=search) |
                             Q(last_name__icontains=search) | Q(email__icontains=search))
    if role_filter:
        users = users.filter(role=role_filter)
    paginator = Paginator(users, 15)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'dashboard/users/list.html', {
        'users': page_obj,
        'page_obj': page_obj,
        'search': search,
        'role_filter': role_filter,
        'role_choices': CustomUser.ROLE_CHOICES,
    })


@admin_required
def user_detail(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    from orders.models import Order
    all_orders = Order.objects.filter(user=user)
    user_orders = all_orders.order_by('-created_at')[:10]
    order_stats = all_orders.aggregate(
        total_orders=Count('id'),
        completed_orders=Count('id', filter=Q(status='delivered')),
        pending_orders=Count('id', filter=Q(status__in=['pending', 'processing', 'confirmed', 'shipped'])),
        total_spent=Sum('total', filter=Q(status='delivered')),
    )
    return render(request, 'dashboard/users/detail.html', {
        'profile_user': user,
        'user_orders': user_orders,
        'order_stats': order_stats,
    })


@admin_required
def user_toggle_active(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        old_active = user.is_active
        user.is_active = not user.is_active
        user.save()
        status = 'activated' if user.is_active else 'deactivated'
        log_activity(request, 'update', f'{status.capitalize()} user "{user.username}"',
                     resource='User', resource_label=user.username,
                     previous_value='Active' if old_active else 'Inactive',
                     new_value='Active' if user.is_active else 'Inactive')
        messages.success(request, f'User {user.username} has been {status}.')
    return redirect('dashboard:user_list')


@admin_required
def user_delete(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if user.pk == request.user.pk:
        messages.error(request, "You can't archive your own account.")
        return redirect('dashboard:user_list')
    if user.is_admin_user():
        remaining_admins = CustomUser.objects.filter(Q(role='admin') | Q(is_superuser=True)).exclude(pk=user.pk).count()
        if remaining_admins == 0:
            messages.error(request, "You can't archive the last remaining admin account.")
            return redirect('dashboard:user_list')
    if request.method == 'POST':
        name = user.username
        role_label = user.get_role_display()
        user.archive(by_user=request.user)
        log_activity(request, 'delete', f'Archived {role_label.lower()} account "{name}"',
                     resource='User', resource_label=name,
                     previous_value=role_label)
        messages.success(request, f'Account "{name}" moved to Archive. You can restore it anytime.')
        return redirect('dashboard:user_list')
    return render(request, 'dashboard/users/confirm_delete.html', {'profile_user': user})


@admin_required
def user_restore(request, pk):
    user = get_object_or_404(CustomUser.all_objects, pk=pk, is_deleted=True)
    if request.method == 'POST':
        user.restore()
        log_activity(request, 'restore', f'Restored account "{user.username}"',
                     resource='User', resource_label=user.username)
        messages.success(request, f'Account "{user.username}" restored.')
    return redirect('dashboard:archive')


# ─── Archive ────────────────────────────────────────────────────────────────────
@admin_required
def archive(request):
    tab = request.GET.get('tab', 'products')
    products = Product.all_objects.filter(is_deleted=True).select_related('deleted_by', 'category').order_by('-deleted_at')
    categories = Category.all_objects.filter(is_deleted=True).select_related('deleted_by').order_by('-deleted_at')
    customers = CustomUser.all_objects.filter(is_deleted=True, role='customer').select_related('deleted_by').order_by('-deleted_at')
    staff = CustomUser.all_objects.filter(is_deleted=True).exclude(role='customer').select_related('deleted_by').order_by('-deleted_at')
    return render(request, 'dashboard/archive/list.html', {
        'tab': tab,
        'products': products,
        'categories': categories,
        'customers': customers,
        'staff': staff,
        'counts': {
            'products': products.count(),
            'categories': categories.count(),
            'customers': customers.count(),
            'staff': staff.count(),
        },
    })


# ─── Activity Log ───────────────────────────────────────────────────────────────
def _format_duration(td):
    """Human-friendly duration string, e.g. '2 min', '1h 14m', '3d 2h'."""
    if td is None:
        return None
    total_seconds = int(td.total_seconds())
    if total_seconds < 60:
        return f'{max(total_seconds, 0)}s'
    minutes = total_seconds // 60
    if minutes < 60:
        return f'{minutes} min'
    hours, minutes = divmod(minutes, 60)
    if hours < 24:
        return f'{hours}h {minutes}m' if minutes else f'{hours}h'
    days, hours = divmod(hours, 24)
    return f'{days}d {hours}h' if hours else f'{days}d'


def _build_activity_sessions(logs):
    """
    Groups a chronological (ascending) iterable of ActivityLog rows into one
    "session" per staff login: a session starts at 'login', ends at the next
    'logout' for that same username, and collects every action performed
    in between. This turns the old one-row-per-event log into one card per
    login session, with the individual actions nested inside it.

    Sessions with no matching login (actions/logout that started before the
    filtered window) and sessions with no matching logout yet (still signed
    in) are both included.
    """
    open_sessions = {}   # username -> in-progress session dict
    sessions = []

    def new_session(log, login_time=None, logout_time=None):
        return {
            'username': log.username,
            'role': log.role,
            'user_id': log.user_id,
            'login_time': login_time,
            'login_log_id': log.id if login_time else None,
            'logout_time': logout_time,
            'logout_log_id': log.id if logout_time else None,
            'ip': log.ip_address,
            'actions': [],
        }

    for log in logs:
        key = log.username or f'user_{log.user_id}'
        if log.action == 'login':
            # Previous session for this user never saw a logout (e.g. browser
            # closed) — close it out as-is before starting the new one.
            if key in open_sessions:
                sessions.append(open_sessions.pop(key))
            open_sessions[key] = new_session(log, login_time=log.timestamp)
        elif log.action == 'logout':
            sess = open_sessions.pop(key, None) or new_session(log)
            sess['logout_time'] = log.timestamp
            sess['logout_log_id'] = log.id
            sess['ip'] = log.ip_address
            sessions.append(sess)
        else:
            sess = open_sessions.setdefault(key, new_session(log))
            sess['actions'].append(log)
            sess['ip'] = log.ip_address

    # Anything still open means that staff member hasn't logged out yet.
    sessions.extend(open_sessions.values())

    for s in sessions:
        s['actions'].sort(key=lambda a: a.timestamp, reverse=True)
        s['action_count'] = len(s['actions'])
        # A session card links to *some* real ActivityLog row for its
        # "view details" button — prefer the login event, then the most
        # recent action, then the logout event.
        s['detail_log_id'] = (
            s['login_log_id'] or (s['actions'][0].id if s['actions'] else None) or s['logout_log_id']
        )
        if s['login_time'] and s['logout_time']:
            s['duration_display'] = _format_duration(s['logout_time'] - s['login_time'])
        elif s['login_time'] and not s['logout_time']:
            s['duration_display'] = _format_duration(timezone.now() - s['login_time'])
        else:
            s['duration_display'] = None
        if s['login_time']:
            s['sort_time'] = s['login_time']
        elif s['actions']:
            s['sort_time'] = s['actions'][-1].timestamp  # actions are desc-sorted; [-1] is earliest
        else:
            s['sort_time'] = s['logout_time']

    sessions.sort(key=lambda s: s['sort_time'], reverse=True)
    return sessions


@admin_required
def activity_log(request):
    STAFF_ROLES = ['admin', 'staff', 'coordinator']

    search = request.GET.get('q', '')
    staff_filter = request.GET.get('staff', '')
    action_filter = request.GET.get('action', '')
    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')

    # This page is the STAFF activity log — customer logins/actions are
    # tracked elsewhere and shouldn't clutter it, so every query below is
    # scoped to staff/coordinator/admin accounts only.
    staff_usernames = set(
        CustomUser.objects.filter(role__in=STAFF_ROLES).values_list('username', flat=True)
    )

    base_qs = ActivityLog.objects.filter(Q(role__in=STAFF_ROLES) | Q(username__in=staff_usernames))
    if date_from:
        base_qs = base_qs.filter(timestamp__date__gte=date_from)
    if date_to:
        base_qs = base_qs.filter(timestamp__date__lte=date_to)

    # Failed logins have no reliable staff account (the username may not
    # even exist), so they're tracked separately rather than folded into a
    # session card. Still restricted to attempts against known staff
    # usernames, to keep this page staff-only.
    failed_logins = base_qs.filter(action='login_failed')
    if search:
        failed_logins = failed_logins.filter(Q(username__icontains=search) | Q(ip_address__icontains=search))
    failed_logins = failed_logins.order_by('-timestamp')[:10]

    session_logs = base_qs.exclude(action='login_failed')
    if staff_filter:
        session_logs = session_logs.filter(username=staff_filter)
    if search:
        session_logs = session_logs.filter(
            Q(username__icontains=search) | Q(description__icontains=search) | Q(ip_address__icontains=search)
        )

    sessions = _build_activity_sessions(session_logs.order_by('timestamp'))

    if action_filter:
        sessions = [s for s in sessions if any(a.action == action_filter for a in s['actions'])]

    # Group sessions by user so the main list shows one card per staff
    # member (not one card per login session) — clicking a card opens
    # that user's full history via activity_log_user.
    user_groups = {}
    for s in sessions:
        key = s['username'] or f"user_{s['user_id']}"
        g = user_groups.setdefault(key, {
            'username': s['username'],
            'role': s['role'],
            'user_id': s['user_id'],
            'session_count': 0,
            'action_count': 0,
            'is_active_now': False,
            'latest_sort_time': s['sort_time'],
            'latest_session': s,
        })
        g['session_count'] += 1
        g['action_count'] += s['action_count']
        if s['logout_time'] is None and s['login_time']:
            g['is_active_now'] = True
        if s['sort_time'] and (not g['latest_sort_time'] or s['sort_time'] > g['latest_sort_time']):
            g['latest_sort_time'] = s['sort_time']
            g['latest_session'] = s

    user_rows = sorted(user_groups.values(), key=lambda g: g['latest_sort_time'] or timezone.now(), reverse=True)

    today = timezone.localdate()
    stats = {
        'total_sessions': len(sessions),
        'sessions_today': sum(1 for s in sessions if s['login_time'] and s['login_time'].date() == today),
        'active_now': sum(1 for s in sessions if s['logout_time'] is None),
        'failed_today': base_qs.filter(action='login_failed', timestamp__date=today).count(),
    }

    staff_choices = list(
        ActivityLog.objects.filter(Q(role__in=STAFF_ROLES) | Q(username__in=staff_usernames))
        .exclude(username='').values_list('username', flat=True).distinct().order_by('username')
    )

    paginator = Paginator(user_rows, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'dashboard/activity/list.html', {
        'user_rows': page_obj,
        'page_obj': page_obj,
        'failed_logins': failed_logins,
        'search': search,
        'staff_filter': staff_filter,
        'action_filter': action_filter,
        'date_from': date_from,
        'date_to': date_to,
        'action_choices': [c for c in ActivityLog.ACTION_CHOICES if c[0] not in ('login', 'logout', 'login_failed')],
        'staff_choices': staff_choices,
        'stats': stats,
    })


@admin_required
def activity_log_export_csv(request):
    STAFF_ROLES = ['admin', 'staff', 'coordinator']
    staff_usernames = set(
        CustomUser.objects.filter(role__in=STAFF_ROLES).values_list('username', flat=True)
    )
    logs = ActivityLog.objects.filter(
        Q(role__in=STAFF_ROLES) | Q(username__in=staff_usernames)
    ).order_by('-timestamp')

    search = request.GET.get('q', '')
    staff_filter = request.GET.get('staff', '')
    action_filter = request.GET.get('action', '')
    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')
    if search:
        logs = logs.filter(Q(username__icontains=search) | Q(description__icontains=search) |
                            Q(ip_address__icontains=search))
    if staff_filter:
        logs = logs.filter(username=staff_filter)
    if action_filter:
        logs = logs.filter(action=action_filter)
    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)
    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)

    response = HttpResponse(content_type='text/csv')
    now = timezone.now()
    response['Content-Disposition'] = f'attachment; filename="Likhalaya_ActivityLog_{now.strftime("%Y%m%d")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Timestamp', 'User', 'Role', 'Action', 'Description', 'Resource', 'Previous Value',
                      'New Value', 'Status', 'IP Address'])
    for log in logs:
        writer.writerow([
            timezone.localtime(log.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
            log.username, log.role, log.get_action_display(), log.description,
            f'{log.resource} — {log.resource_label}'.strip(' —') if (log.resource or log.resource_label) else '',
            log.previous_value, log.new_value, log.get_status_display(), log.ip_address or '',
        ])
    return response


@admin_required
def activity_log_detail(request, pk):
    STAFF_ROLES = ['admin', 'staff', 'coordinator']
    staff_usernames = set(
        CustomUser.objects.filter(role__in=STAFF_ROLES).values_list('username', flat=True)
    )
    log = get_object_or_404(
        ActivityLog.objects.filter(Q(role__in=STAFF_ROLES) | Q(username__in=staff_usernames)),
        pk=pk,
    )

    # Multi-field updates (e.g. category/product edits) are stored as
    # "Field: old → new; Field2: old2 → new2" in previous_value. Split
    # that into individual rows here so the template can render a clean
    # list instead of one long wrapped string.
    field_changes = []
    if log.previous_value and ': ' in log.previous_value and ' → ' in log.previous_value:
        parts = [p.strip() for p in log.previous_value.split('; ') if p.strip()]
        all_parsed = True
        for part in parts:
            if ': ' not in part or ' → ' not in part:
                all_parsed = False
                break
            label, rest = part.split(': ', 1)
            old_val, new_val = rest.split(' → ', 1)
            field_changes.append({'label': label, 'old': old_val, 'new': new_val})
        if not all_parsed:
            field_changes = []

    # For image fields, resolve the stored filenames into actual media
    # URLs (based on which resource logged the change) so the detail
    # page can render thumbnails instead of just the filename text.
    image_folder = {'Product': 'products/', 'Category': 'categories/'}.get(log.resource)
    if image_folder:
        for change in field_changes:
            if change['label'] == 'Image':
                if change['old'] and change['old'] != '(none)':
                    change['old_url'] = settings.MEDIA_URL + image_folder + change['old']
                if change['new'] and change['new'] != '(none)':
                    change['new_url'] = settings.MEDIA_URL + image_folder + change['new']

    return render(request, 'dashboard/activity/detail.html', {'log': log, 'field_changes': field_changes})


@admin_required
def activity_log_user(request, username):
    """
    Full activity history for a single staff member, reached by clicking
    their name on the main Activity Log page (opens in a new window).
    Every row here links through to activity_log_detail for that entry.
    """
    STAFF_ROLES = ['admin', 'staff', 'coordinator']
    staff_usernames = set(
        CustomUser.objects.filter(role__in=STAFF_ROLES).values_list('username', flat=True)
    )
    all_logs = ActivityLog.objects.filter(
        Q(role__in=STAFF_ROLES) | Q(username__in=staff_usernames),
        username=username,
    ).order_by('-timestamp')

    latest = all_logs.first()

    logs = all_logs
    search = request.GET.get('q', '')
    action_filter = request.GET.get('action', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('from', '')
    date_to = request.GET.get('to', '')

    if search:
        logs = logs.filter(Q(description__icontains=search) | Q(resource_label__icontains=search) | Q(resource__icontains=search))
    if action_filter:
        logs = logs.filter(action=action_filter)
    if status_filter:
        logs = logs.filter(status=status_filter)
    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)
    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)

    stats = all_logs.aggregate(
        total=Count('id'),
        successful=Count('id', filter=Q(status='success')),
        failed=Count('id', filter=Q(status='failed')),
        logins=Count('id', filter=Q(action='login')),
        updates=Count('id', filter=Q(action='update')),
        orders_processed=Count('id', filter=Q(resource='Order')),
    )

    paginator = Paginator(logs, 15)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Day-group labels ("Today" / "Yesterday" / date) for the timeline view —
    # computed here since Django templates can't do "yesterday" comparisons.
    today = timezone.localdate()
    yesterday = today - timedelta(days=1)
    for log in page_obj:
        log_date = timezone.localtime(log.timestamp).date()
        if log_date == today:
            log.day_label = 'Today'
        elif log_date == yesterday:
            log.day_label = 'Yesterday'
        else:
            log.day_label = log_date.strftime('%B %d, %Y')

    return render(request, 'dashboard/activity/user_list.html', {
        'username': username,
        'role': latest.role if latest else '',
        'last_active': latest.timestamp if latest else None,
        'logs': page_obj,
        'page_obj': page_obj,
        'search': search,
        'action_filter': action_filter,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'action_choices': [c for c in ActivityLog.ACTION_CHOICES if c[0] != 'login_failed'],
        'status_choices': ActivityLog.STATUS_CHOICES,
        'total_count': logs.count(),
        'stats': stats,
    })


# ─── Categories ─────────────────────────────────────────────────────────────────
@staff_required
def category_list(request):
    cats = Category.objects.annotate(product_count=Count('products')).order_by('order', 'name')
    return render(request, 'dashboard/categories/list.html', {'categories': cats})


@admin_required
def category_create(request):
    from store.forms import CategoryForm
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            cat = form.save()
            log_activity(request, 'create', f'Created category "{cat.name}"',
                         resource='Category', resource_label=cat.name)
            messages.success(request, 'Category created.')
            return redirect('dashboard:category_list')
    else:
        form = CategoryForm()
    return render(request, 'dashboard/categories/form.html', {'form': form, 'title': 'Add Category'})


@admin_required
def category_edit(request, pk):
    from store.forms import CategoryForm
    cat = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        # Snapshot every trackable field before the form overwrites them,
        # so the activity log can report whichever ones actually changed
        # (name, description, order, active status, or the image).
        old_values = {
            'name': cat.name,
            'description': cat.description,
            'order': cat.order,
            'is_active': cat.is_active,
            'image': cat.image.name if cat.image else '',
        }
        form = CategoryForm(request.POST, request.FILES, instance=cat)
        if form.is_valid():
            form.save()

            new_values = {
                'name': cat.name,
                'description': cat.description,
                'order': cat.order,
                'is_active': cat.is_active,
                'image': cat.image.name if cat.image else '',
            }

            field_labels = {
                'name': 'Name',
                'description': 'Description',
                'order': 'Order',
                'is_active': 'Active status',
                'image': 'Image',
            }

            changes = []
            for field, old_val in old_values.items():
                new_val = new_values[field]
                if old_val != new_val:
                    if field == 'image':
                        old_display = old_val.rsplit('/', 1)[-1] if old_val else '(none)'
                        new_display = new_val.rsplit('/', 1)[-1] if new_val else '(none)'
                    elif field == 'description':
                        old_display = old_val or '(empty)'
                        new_display = new_val or '(empty)'
                    elif field == 'is_active':
                        old_display = 'On' if old_val else 'Off'
                        new_display = 'On' if new_val else 'Off'
                    else:
                        old_display = old_val
                        new_display = new_val
                    changes.append(f'{field_labels[field]}: {old_display} → {new_display}')

            previous_value = '; '.join(changes) if changes else 'No fields changed'
            new_value = f'{len(changes)} field(s) updated' if changes else 'No fields changed'

            log_activity(request, 'update', f'Updated category "{cat.name}"',
                         resource='Category', resource_label=cat.name,
                         previous_value=previous_value, new_value=new_value)
            messages.success(request, 'Category updated.')
            return redirect('dashboard:category_list')
    else:
        form = CategoryForm(instance=cat)
    return render(request, 'dashboard/categories/form.html', {'form': form, 'title': 'Edit Category', 'category': cat})


@admin_required
def category_delete(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        cat_name = cat.name
        cat.archive(by_user=request.user)
        log_activity(request, 'delete', f'Archived category "{cat_name}"',
                     resource='Category', resource_label=cat_name,
                     previous_value=cat_name)
        messages.success(request, f'Category "{cat_name}" moved to Archive. You can restore it anytime.')
        return redirect('dashboard:category_list')
    return render(request, 'dashboard/categories/confirm_delete.html', {'category': cat})


@admin_required
def category_restore(request, pk):
    cat = get_object_or_404(Category.all_objects, pk=pk, is_deleted=True)
    if request.method == 'POST':
        cat.restore()
        log_activity(request, 'restore', f'Restored category "{cat.name}"',
                     resource='Category', resource_label=cat.name)
        messages.success(request, f'Category "{cat.name}" restored.')
    return redirect('dashboard:archive')


# ─── Reports ────────────────────────────────────────────────────────────────────
@staff_required
def reports(request):
    now = timezone.now()
    thirty_days_ago = now - timedelta(days=30)

    # ── Calendar-month period filter ──────────────────────────────────────────
    # 'period' picks WHICH slice of the existing Order/OrderItem history to
    # report on. Nothing is ever deleted for reporting — we just narrow the
    # date range of the same query. On the 1st of a new month, 'this_month'
    # naturally starts from ₱0 because no orders exist yet with that timestamp;
    # last month's data is still fully intact and reachable via 'last_month'
    # or 'all'.
    period = request.GET.get('period', 'all')

    this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_end = this_month_start - timedelta(microseconds=1)
    last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    orders_qs = Order.objects.all()
    items_qs = OrderItem.objects.all()
    period_label = 'All Time'

    if period == 'this_month':
        orders_qs = orders_qs.filter(created_at__gte=this_month_start)
        items_qs = items_qs.filter(order__created_at__gte=this_month_start)
        period_label = now.strftime('%B %Y')
    elif period == 'last_month':
        orders_qs = orders_qs.filter(created_at__gte=last_month_start, created_at__lt=this_month_start)
        items_qs = items_qs.filter(order__created_at__gte=last_month_start, order__created_at__lt=this_month_start)
        period_label = last_month_start.strftime('%B %Y')

    # Sales by status (respects the period filter above)
    sales_by_status = orders_qs.values('status').annotate(
        count=Count('id'), revenue=Sum('total')
    ).order_by('-count')

    # Top products by units sold, within the selected period
    top_products = (items_qs.values('product_name')
                    .annotate(units=Sum('quantity'), revenue=Sum('product_price'))
                    .order_by('-units')[:10])

    # Monthly trend (always shows the last 180 days regardless of period,
    # so the chart gives context even when a single month is selected)
    monthly_trend = (Order.objects.filter(created_at__gte=now - timedelta(days=180))
                     .annotate(month=TruncMonth('created_at'))
                     .values('month')
                     .annotate(orders=Count('id'), revenue=Sum('total'))
                     .order_by('month'))

    ctx = {
        'sales_by_status': sales_by_status,
        'top_products': top_products,
        'monthly_trend': monthly_trend,
        'total_revenue': orders_qs.filter(status='delivered').aggregate(t=Sum('total'))['t'] or 0,
        'total_orders': orders_qs.count(),
        'avg_order': orders_qs.aggregate(a=Avg('total'))['a'] or 0,
        'period': period,
        'period_label': period_label,
    }
    return render(request, 'dashboard/reports.html', ctx)


@admin_required
def reports_export_excel(request):
    """Branded, print-ready Excel version of the Sales Reports page."""
    now = timezone.now()

    # Same period filter as the on-screen report, so what staff sees is what
    # they export. Defaults to 'all' if no period is passed.
    period = request.GET.get('period', 'all')

    this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_end = this_month_start - timedelta(microseconds=1)
    last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    orders_qs = Order.objects.all()
    items_qs = OrderItem.objects.all()
    period_label = 'All Time'

    if period == 'this_month':
        orders_qs = orders_qs.filter(created_at__gte=this_month_start)
        items_qs = items_qs.filter(order__created_at__gte=this_month_start)
        period_label = now.strftime('%B %Y')
    elif period == 'last_month':
        orders_qs = orders_qs.filter(created_at__gte=last_month_start, created_at__lt=this_month_start)
        items_qs = items_qs.filter(order__created_at__gte=last_month_start, order__created_at__lt=this_month_start)
        period_label = last_month_start.strftime('%B %Y')

    sales_by_status = orders_qs.values('status').annotate(
        count=Count('id'), revenue=Sum('total')
    ).order_by('-count')

    top_products = (items_qs.values('product_name')
                    .annotate(units=Sum('quantity'), revenue=Sum('product_price'))
                    .order_by('-units')[:10])

    total_revenue = orders_qs.filter(status='delivered').aggregate(t=Sum('total'))['t'] or 0
    total_orders = orders_qs.count()
    avg_order = orders_qs.aggregate(a=Avg('total'))['a'] or 0

    # Per-month breakdown for the detailed report. Always covers the last 12
    # months regardless of the on-screen period filter, so the export gives
    # a fuller month-by-month picture without needing to re-export per month.
    twelve_months_ago = (this_month_start - timedelta(days=365)).replace(day=1)
    monthly_breakdown = (Order.objects.filter(created_at__gte=twelve_months_ago)
                         .annotate(month=TruncMonth('created_at'))
                         .values('month')
                         .annotate(orders=Count('id'), revenue=Sum('total'))
                         .order_by('month'))

    # Top products sold per month (last 12 months). Grouped in Python after
    # one query, so it's a single DB round-trip: for each month, keep the top
    # 5 products by units sold, highest first.
    monthly_product_rows = (OrderItem.objects.filter(order__created_at__gte=twelve_months_ago)
                            .annotate(month=TruncMonth('order__created_at'))
                            .values('month', 'product_name')
                            .annotate(units=Sum('quantity'), revenue=Sum('product_price'))
                            .order_by('month', '-units'))

    monthly_top_products = {}
    for r in monthly_product_rows:
        monthly_top_products.setdefault(r['month'], []).append(r)
    for month_key in monthly_top_products:
        monthly_top_products[month_key] = monthly_top_products[month_key][:5]

    status_display = dict(Order.STATUS_CHOICES)

    # ── Simple, easy-to-read styling: black text, white/light-gray fills only.
    # No bright colors — just bold/plain and light shading to separate sections.
    BLACK = '000000'
    DARK_GRAY = '333333'
    MID_GRAY = '666666'
    LIGHT_GRAY = 'F2F2F2'
    HEADER_GRAY = 'D9D9D9'
    WHITE = 'FFFFFF'

    title_font = Font(name='Calibri', size=18, bold=True, color=BLACK)
    subtitle_font = Font(name='Calibri', size=10, italic=True, color=MID_GRAY)
    section_font = Font(name='Calibri', size=12, bold=True, color=BLACK)
    header_font = Font(name='Calibri', size=10, bold=True, color=BLACK)
    label_font = Font(name='Calibri', size=10, color=MID_GRAY)
    value_font = Font(name='Calibri', size=16, bold=True, color=BLACK)
    body_font = Font(name='Calibri', size=10, color=DARK_GRAY)
    total_font = Font(name='Calibri', size=10, bold=True, color=BLACK)

    section_fill = PatternFill('solid', fgColor=HEADER_GRAY)
    header_fill = PatternFill('solid', fgColor=HEADER_GRAY)
    card_fill = PatternFill('solid', fgColor=LIGHT_GRAY)
    zebra_fill = PatternFill('solid', fgColor=LIGHT_GRAY)
    total_fill = PatternFill('solid', fgColor=HEADER_GRAY)

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sales Report'
    ws.sheet_view.showGridLines = False

    # ── Column widths ──
    widths = [30, 16, 16, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Letterhead ──
    ws.merge_cells(f'A{row}:E{row}')
    c = ws[f'A{row}']
    c.value = 'LIKHALAYA'
    c.font = title_font
    c.alignment = center
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    c = ws[f'A{row}']
    c.value = 'PDL Market — Sales Report'
    c.font = Font(name='Calibri', size=11, bold=True, color=DARK_GRAY)
    c.alignment = center
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    c = ws[f'A{row}']
    c.value = f'Period: {period_label}'
    c.font = Font(name='Calibri', size=10, bold=True, color=BLACK)
    c.alignment = center
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    c = ws[f'A{row}']
    c.value = f'Generated {now.strftime("%B %d, %Y at %I:%M %p")}'
    c.font = subtitle_font
    c.alignment = center
    row += 2

    # ── Summary cards ──
    card_specs = [
        ('Total Revenue', f'₱{total_revenue:,.0f}'),
        ('Total Orders', f'{total_orders}'),
        ('Avg Order Value', f'₱{avg_order:,.0f}'),
    ]
    card_cols = [('A', 'B'), ('C', 'C'), ('D', 'E')]
    label_row, value_row = row, row + 1
    for (label, value), (c1, c2) in zip(card_specs, card_cols):
        rng_label = f'{c1}{label_row}:{c2}{label_row}'
        rng_value = f'{c1}{value_row}:{c2}{value_row}'
        ws.merge_cells(rng_label)
        ws.merge_cells(rng_value)
        lc = ws[f'{c1}{label_row}']
        lc.value = label
        lc.font = label_font
        lc.alignment = center
        lc.fill = card_fill
        vc = ws[f'{c1}{value_row}']
        vc.value = value
        vc.font = value_font
        vc.alignment = center
        vc.fill = card_fill
        for rr in (label_row, value_row):
            for col_letter in (c1, c2):
                ws[f'{col_letter}{rr}'].border = border
    ws.row_dimensions[label_row].height = 18
    ws.row_dimensions[value_row].height = 28
    row = value_row + 2

    # ── Orders by Status ──
    ws.merge_cells(f'A{row}:E{row}')
    c = ws[f'A{row}']
    c.value = 'Orders by Status'
    c.font = section_font
    c.fill = section_fill
    c.alignment = center
    ws.row_dimensions[row].height = 20
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col_letter}{row}'].fill = section_fill
    row += 1

    headers = ['Status', 'Orders', 'Revenue']
    header_cells = ['A', 'C', 'E']
    ws.merge_cells(f'A{row}:B{row}')
    ws.merge_cells(f'C{row}:D{row}')
    for h, col in zip(headers, header_cells):
        cell = ws[f'{col}{row}']
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col_letter}{row}'].fill = header_fill
        ws[f'{col_letter}{row}'].border = border
    row += 1

    status_start = row
    total_status_orders = 0
    total_status_revenue = 0
    for i, s in enumerate(sales_by_status):
        label = status_display.get(s['status'], s['status'])
        count = s['count']
        revenue = s['revenue'] or 0
        total_status_orders += count
        total_status_revenue += revenue
        fill = zebra_fill if i % 2 else PatternFill('solid', fgColor=WHITE)

        ws.merge_cells(f'A{row}:B{row}')
        badge = ws[f'A{row}']
        badge.value = label.upper()
        badge.font = Font(name='Calibri', size=9, bold=True, color=BLACK)
        badge.fill = fill
        badge.alignment = center
        badge.border = border

        ws.merge_cells(f'C{row}:D{row}')
        oc = ws[f'C{row}']
        oc.value = count
        oc.font = body_font
        oc.alignment = center
        oc.fill = fill
        oc.border = border

        rv = ws[f'E{row}']
        rv.value = f'₱{revenue:,.0f}'
        rv.font = body_font
        rv.alignment = center
        rv.fill = fill
        rv.border = border
        ws[f'B{row}'].fill = fill
        ws[f'B{row}'].border = border
        row += 1

    # Totals row
    ws.merge_cells(f'A{row}:B{row}')
    tc = ws[f'A{row}']
    tc.value = 'TOTAL'
    tc.font = total_font
    tc.fill = total_fill
    tc.alignment = center
    tc.border = border
    ws.merge_cells(f'C{row}:D{row}')
    to = ws[f'C{row}']
    to.value = total_status_orders
    to.font = total_font
    to.fill = total_fill
    to.alignment = center
    to.border = border
    tr = ws[f'E{row}']
    tr.value = f'₱{total_status_revenue:,.0f}'
    tr.font = total_font
    tr.fill = total_fill
    tr.alignment = center
    tr.border = border
    ws[f'B{row}'].fill = total_fill
    ws[f'B{row}'].border = border
    row += 3

    # ── Top Products by Units Sold ──
    ws.merge_cells(f'A{row}:E{row}')
    c = ws[f'A{row}']
    c.value = 'Top Products by Units Sold'
    c.font = section_font
    c.alignment = center
    ws.row_dimensions[row].height = 20
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col_letter}{row}'].fill = section_fill
    row += 1

    headers = ['Product', 'Units Sold', 'Revenue']
    ws.merge_cells(f'A{row}:B{row}')
    ws.merge_cells(f'C{row}:D{row}')
    for h, col in zip(headers, header_cells):
        cell = ws[f'{col}{row}']
        cell.value = h
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col_letter}{row}'].fill = header_fill
        ws[f'{col_letter}{row}'].border = border
    row += 1

    for i, p in enumerate(top_products):
        fill = zebra_fill if i % 2 else PatternFill('solid', fgColor=WHITE)
        ws.merge_cells(f'A{row}:B{row}')
        nc = ws[f'A{row}']
        nc.value = p['product_name']
        nc.font = body_font
        nc.alignment = center
        nc.fill = fill
        nc.border = border

        ws.merge_cells(f'C{row}:D{row}')
        uc = ws[f'C{row}']
        uc.value = p['units'] or 0
        uc.font = body_font
        uc.alignment = center
        uc.fill = fill
        uc.border = border

        rv = ws[f'E{row}']
        rv.value = f"₱{(p['revenue'] or 0):,.0f}"
        rv.font = body_font
        rv.alignment = center
        rv.fill = fill
        rv.border = border
        ws[f'B{row}'].fill = fill
        ws[f'B{row}'].border = border
        row += 1

    row += 1

    # ── Monthly Breakdown (last 12 months) ──
    ws.merge_cells(f'A{row}:E{row}')
    c = ws[f'A{row}']
    c.value = 'Monthly Breakdown (Last 12 Months)'
    c.font = section_font
    c.alignment = center
    ws.row_dimensions[row].height = 20
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col_letter}{row}'].fill = section_fill
    row += 1

    headers = ['Month', 'Orders', 'Revenue']
    ws.merge_cells(f'A{row}:B{row}')
    ws.merge_cells(f'C{row}:D{row}')
    for h, col in zip(headers, header_cells):
        cell = ws[f'{col}{row}']
        cell.value = h
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col_letter}{row}'].fill = header_fill
        ws[f'{col_letter}{row}'].border = border
    row += 1

    for i, m in enumerate(monthly_breakdown):
        fill = zebra_fill if i % 2 else PatternFill('solid', fgColor=WHITE)
        ws.merge_cells(f'A{row}:B{row}')
        mc = ws[f'A{row}']
        mc.value = m['month'].strftime('%B %Y') if m['month'] else '—'
        mc.font = body_font
        mc.alignment = center
        mc.fill = fill
        mc.border = border

        ws.merge_cells(f'C{row}:D{row}')
        oc = ws[f'C{row}']
        oc.value = m['orders'] or 0
        oc.font = body_font
        oc.alignment = center
        oc.fill = fill
        oc.border = border

        rv = ws[f'E{row}']
        rv.value = f"₱{(m['revenue'] or 0):,.0f}"
        rv.font = body_font
        rv.alignment = center
        rv.fill = fill
        rv.border = border
        ws[f'B{row}'].fill = fill
        ws[f'B{row}'].border = border
        row += 1

    row += 1

    # ── Top Products per Month (last 12 months) ──
    ws.merge_cells(f'A{row}:E{row}')
    c = ws[f'A{row}']
    c.value = 'Top Products by Units Sold — Per Month (Last 12 Months)'
    c.font = section_font
    c.alignment = center
    ws.row_dimensions[row].height = 20
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col_letter}{row}'].fill = section_fill
    row += 1

    for month_key in sorted(monthly_top_products.keys()):
        products = monthly_top_products[month_key]
        if not products:
            continue

        # Month sub-header row
        ws.merge_cells(f'A{row}:E{row}')
        mh = ws[f'A{row}']
        mh.value = month_key.strftime('%B %Y')
        mh.font = Font(name='Calibri', size=10, bold=True, color=BLACK)
        mh.fill = PatternFill('solid', fgColor=HEADER_GRAY)
        mh.alignment = left
        mh.border = border
        row += 1

        # Column headers for this month's product table
        headers = ['Product', 'Units Sold', 'Revenue']
        ws.merge_cells(f'A{row}:B{row}')
        ws.merge_cells(f'C{row}:D{row}')
        for h, col in zip(headers, header_cells):
            cell = ws[f'{col}{row}']
            cell.value = h
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col_letter}{row}'].border = border
        row += 1

        for i, p in enumerate(products):
            fill = zebra_fill if i % 2 else PatternFill('solid', fgColor=WHITE)
            ws.merge_cells(f'A{row}:B{row}')
            nc = ws[f'A{row}']
            nc.value = p['product_name']
            nc.font = body_font
            nc.alignment = center
            nc.fill = fill
            nc.border = border

            ws.merge_cells(f'C{row}:D{row}')
            uc = ws[f'C{row}']
            uc.value = p['units'] or 0
            uc.font = body_font
            uc.alignment = center
            uc.fill = fill
            uc.border = border

            rv = ws[f'E{row}']
            rv.value = f"₱{(p['revenue'] or 0):,.0f}"
            rv.font = body_font
            rv.alignment = center
            rv.fill = fill
            rv.border = border
            ws[f'B{row}'].fill = fill
            ws[f'B{row}'].border = border
            row += 1

        row += 1  # spacing between months

    ws.merge_cells(f'A{row}:E{row}')
    footer = ws[f'A{row}']
    footer.value = 'Likhalaya PDL Market — Confidential internal sales report'
    footer.font = Font(name='Calibri', size=8, italic=True, color='AAAAAA')
    footer.alignment = center

    # ── Print setup ──
    ws.print_area = f'A1:E{row}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    buffer_bytes = _xlsx_bytes(wb)
    response = HttpResponse(
        buffer_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Likhalaya_Sales_Report_{period}_{now.strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Customer Dashboard ─────────────────────────────────────────────────────────
@login_required
def customer_dashboard(request):
    orders = Order.objects.filter(user=request.user).prefetch_related('items').order_by('-created_at')
    stats = {
        'total_orders': orders.count(),
        'pending': orders.filter(status='pending').count(),
        'delivered': orders.filter(status='delivered').count(),
        'total_spent': orders.filter(status='delivered').aggregate(t=Sum('total'))['t'] or 0,
    }
    recent_orders = orders[:5]
    return render(request, 'dashboard/customer/home.html', {
        'stats': stats,
        'recent_orders': recent_orders,
    })


# ─── Livelihood Videos ──────────────────────────────────────────────────────────
@staff_required
def video_list(request):
    videos = LivelihoodVideo.objects.all().order_by('order', '-created_at')
    return render(request, 'dashboard/videos/list.html', {'videos': videos})


@admin_required
def video_create(request):
    from store.forms import LivelihoodVideoForm
    if request.method == 'POST':
        form = LivelihoodVideoForm(request.POST, request.FILES)
        if form.is_valid():
            video = form.save()
            log_activity(request, 'create', f'Added livelihood video "{video.title}"',
                         resource='LivelihoodVideo', resource_label=video.title)
            messages.success(request, 'Video added.')
            return redirect('dashboard:video_list')
    else:
        form = LivelihoodVideoForm()
    return render(request, 'dashboard/videos/form.html', {'form': form, 'title': 'Add Video'})


@admin_required
def video_edit(request, pk):
    from store.forms import LivelihoodVideoForm
    video = get_object_or_404(LivelihoodVideo, pk=pk)
    if request.method == 'POST':
        form = LivelihoodVideoForm(request.POST, request.FILES, instance=video)
        if form.is_valid():
            form.save()
            log_activity(request, 'update', f'Updated livelihood video "{video.title}"',
                         resource='LivelihoodVideo', resource_label=video.title)
            messages.success(request, 'Video updated.')
            return redirect('dashboard:video_list')
    else:
        form = LivelihoodVideoForm(instance=video)
    return render(request, 'dashboard/videos/form.html', {'form': form, 'title': 'Edit Video', 'video': video})


@admin_required
def video_delete(request, pk):
    video = get_object_or_404(LivelihoodVideo, pk=pk)
    if request.method == 'POST':
        title = video.title
        video.delete()
        log_activity(request, 'delete', f'Deleted livelihood video "{title}"',
                     resource='LivelihoodVideo', resource_label=title)
        messages.success(request, f'Video "{title}" deleted.')
        return redirect('dashboard:video_list')
    return render(request, 'dashboard/videos/confirm_delete.html', {'video': video})